from dataclasses import dataclass, field

from django.contrib.auth.hashers import make_password
from openpyxl import load_workbook

from django.db.models import Count

from .models import Course, User


MAX_IMPORT_ROWS = 1000


@dataclass
class ImportPreview:
    rows: list[dict]
    created: int
    updated: int
    errors: list[str] = field(default_factory=list)


def _cell(row, index):
    value = row[index].value if len(row) > index else None
    return str(value).strip() if value is not None else ""


def _workbook_rows(uploaded_file):
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    return list(workbook.active.iter_rows())


def parse_course_import(uploaded_file) -> ImportPreview:
    records, codes, errors = [], set(), []
    for number, row in enumerate(_workbook_rows(uploaded_file)[1:], start=2):
        code, name = _cell(row, 0), _cell(row, 1)
        if not code and not name:
            continue
        if not code or not name:
            errors.append(f"第 {number} 行缺少课程编号或名称")
            continue
        if code in codes:
            errors.append(f"第 {number} 行课程编号重复：{code}")
            continue
        try:
            capacity = int(_cell(row, 3) or 1)
        except (TypeError, ValueError):
            errors.append(f"第 {number} 行名额必须是整数")
            continue
        if capacity < 1:
            errors.append(f"第 {number} 行名额必须大于 0")
            continue
        codes.add(code)
        records.append({"code": code, "name": name, "category": _cell(row, 2), "capacity": capacity,
                        "description": _cell(row, 4)})
    if not records and not errors:
        raise ValueError("文件中没有可导入的数据")
    if len(records) + len(errors) > MAX_IMPORT_ROWS:
        raise ValueError(f"一次最多导入 {MAX_IMPORT_ROWS} 条数据")
    existing = set(Course.objects.filter(code__in=codes).values_list("code", flat=True))
    selected_counts = dict(
        Course.objects.filter(code__in=codes).annotate(total=Count("selections")).values_list("code", "total")
    )
    valid_records = []
    for row in records:
        if row["capacity"] < selected_counts.get(row["code"], 0):
            errors.append(f"课程 {row['code']} 的名额不能低于当前已选教师人数")
        else:
            valid_records.append(row)
    return ImportPreview(valid_records, sum(row["code"] not in existing for row in valid_records),
                         sum(row["code"] in existing for row in valid_records), errors)


def parse_teacher_import(uploaded_file) -> ImportPreview:
    records, usernames, errors = [], set(), []
    for number, row in enumerate(_workbook_rows(uploaded_file)[1:], start=2):
        username, name, department, password = (_cell(row, i) for i in range(4))
        if not any((username, name, department, password)):
            continue
        if not username or not name:
            errors.append(f"第 {number} 行缺少登录账号或姓名")
            continue
        if username in usernames:
            errors.append(f"第 {number} 行登录账号重复：{username}")
            continue
        if password and len(password) < 8:
            errors.append(f"第 {number} 行密码不少于 8 位")
            continue
        usernames.add(username)
        records.append({"username": username, "display_name": name, "department": department,
                        "password_hash": make_password(password) if password else ""})
    if not records and not errors:
        raise ValueError("文件中没有可导入的数据")
    if len(records) + len(errors) > MAX_IMPORT_ROWS:
        raise ValueError(f"一次最多导入 {MAX_IMPORT_ROWS} 条数据")
    existing = set(User.objects.filter(username__in=usernames).values_list("username", flat=True))
    admins = set(User.objects.filter(username__in=usernames, role=User.Role.ADMIN).values_list("username", flat=True))
    valid_records = []
    for row in records:
        if row["username"] in admins:
            errors.append(f"账号与管理员账号冲突：{row['username']}")
        elif row["username"] not in existing and not row["password_hash"]:
            errors.append(f"新账号必须填写初始密码：{row['username']}")
        else:
            valid_records.append(row)
    return ImportPreview(valid_records, sum(row["username"] not in existing for row in valid_records),
                         sum(row["username"] in existing for row in valid_records), errors)
