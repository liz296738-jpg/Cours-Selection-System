from io import BytesIO

from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import transaction
from django.core.paginator import Paginator
from django.db.models import Count, Prefetch, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .decorators import portal_admin_required, teacher_required
from .forms import CourseForm, ExcelUploadForm, SiteSettingForm, StyledPasswordChangeForm, TeacherForm
from .imports import parse_course_import, parse_teacher_import
from .models import AuditLog, Course, Selection, SiteSetting, User
from .services import cancel_selection, select_course


def home(request):
    if not request.user.is_authenticated:
        return redirect("login")
    return redirect("dashboard")


@login_required
def dashboard(request):
    if request.user.must_change_password and not request.user.is_portal_admin:
        return redirect("change_password")
    return redirect("control_dashboard" if request.user.is_portal_admin else "teacher_courses")


@login_required
def change_password(request):
    form = StyledPasswordChangeForm(request.user, request.POST or None)
    if request.method == "POST" and form.is_valid():
        user = form.save()
        user.must_change_password = False
        user.save(update_fields=["must_change_password"])
        update_session_auth_hash(request, user)
        AuditLog.objects.create(user=user, action="修改密码")
        messages.success(request, "密码修改成功。")
        return redirect("dashboard")
    return render(request, "registration/change_password.html", {"form": form})


@portal_admin_required
def control_dashboard(request):
    setting = SiteSetting.load()
    context = {
        "setting": setting,
        "course_count": Course.objects.count(),
        "active_course_count": Course.objects.filter(is_active=True).count(),
        "teacher_count": User.objects.filter(role=User.Role.TEACHER, is_active=True).count(),
        "selection_count": Selection.objects.count(),
        "recent_logs": AuditLog.objects.select_related("user")[:8],
    }
    return render(request, "selection/control_dashboard.html", context)


@portal_admin_required
def settings_edit(request):
    setting = SiteSetting.load()
    form = SiteSettingForm(request.POST or None, instance=setting)
    if request.method == "POST" and form.is_valid():
        form.save()
        AuditLog.objects.create(user=request.user, action="修改选课设置",
                                detail=f"每人最多 {setting.max_courses_per_teacher} 门")
        messages.success(request, "选课设置已保存。")
        return redirect("settings_edit")
    return render(request, "selection/settings_form.html", {"form": form, "setting": setting})


@portal_admin_required
def course_list(request):
    query = request.GET.get("q", "").strip()
    courses = Course.objects.annotate(selection_total=Count("selections")).order_by("code", "id")
    if query:
        courses = courses.filter(Q(code__icontains=query) | Q(name__icontains=query) | Q(category__icontains=query))
    return render(request, "selection/course_list.html", {
        "courses": _paginate(request, courses), "query": query, "page_size": _page_size(request),
    })


@portal_admin_required
def course_edit(request, pk=None):
    course = get_object_or_404(Course, pk=pk) if pk else None
    form = CourseForm(request.POST or None, instance=course)
    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        if not obj.created_by_id:
            obj.created_by = request.user
        obj.save()
        AuditLog.objects.create(user=request.user, action="编辑课程" if course else "新增课程", detail=str(obj))
        messages.success(request, "课程已保存。")
        return redirect("course_list")
    return render(request, "selection/course_form.html", {"form": form, "course": course})


@portal_admin_required
@require_POST
def course_toggle(request, pk):
    course = get_object_or_404(Course, pk=pk)
    course.is_active = not course.is_active
    course.save(update_fields=["is_active"])
    AuditLog.objects.create(user=request.user, action="调整课程状态", detail=f"{course}：{'可选' if course.is_active else '停用'}")
    messages.success(request, "课程状态已更新。")
    return redirect("course_list")


@portal_admin_required
def teacher_list(request):
    teachers = User.objects.filter(role=User.Role.TEACHER).annotate(selection_total=Count("selections")).order_by("username", "id")
    query = request.GET.get("q", "").strip()
    if query:
        teachers = teachers.filter(
            Q(username__icontains=query) | Q(display_name__icontains=query) | Q(department__icontains=query)
        )
    return render(request, "selection/teacher_list.html", {
        "teachers": _paginate(request, teachers), "query": query, "page_size": _page_size(request),
    })


@portal_admin_required
def teacher_edit(request, pk=None):
    teacher = get_object_or_404(User, pk=pk, role=User.Role.TEACHER) if pk else None
    form = TeacherForm(request.POST or None, instance=teacher)
    if request.method == "POST" and form.is_valid():
        obj = form.save()
        AuditLog.objects.create(user=request.user, action="编辑教师" if teacher else "新增教师", detail=str(obj))
        messages.success(request, "教师账号已保存。")
        return redirect("teacher_list")
    return render(request, "selection/teacher_form.html", {
        "form": form, "teacher": teacher,
        "teacher_selection_total": teacher.selections.count() if teacher else 0,
    })


def _page_size(request):
    try:
        value = int(request.GET.get("page_size", 20))
    except (TypeError, ValueError):
        value = 20
    return value if value in {20, 50, 100} else 20


def _paginate(request, queryset):
    return Paginator(queryset, _page_size(request)).get_page(request.GET.get("page"))


def _pending_import_key(kind):
    return f"pending_import_{kind}"


def _preview_import(request, kind, parser):
    form = ExcelUploadForm(request.POST or None, request.FILES or None)
    preview = None
    if request.method == "POST" and request.POST.get("action") == "preview" and form.is_valid():
        try:
            parsed = parser(form.cleaned_data["file"])
            request.session[_pending_import_key(kind)] = {"rows": parsed.rows}
            request.session.modified = True
            preview = parsed
        except ValueError as exc:
            form.add_error("file", str(exc))
    return form, preview


@portal_admin_required
def import_courses(request):
    if request.method == "POST" and request.POST.get("action") == "confirm":
        pending = request.session.get(_pending_import_key("courses"))
        if not pending:
            messages.error(request, "导入预览已失效，请重新选择文件。")
        else:
            created = updated = 0
            try:
                with transaction.atomic():
                    for row in pending["rows"]:
                        course = Course.objects.select_for_update().filter(code=row["code"]).first()
                        if course:
                            if course.selections.count() > row["capacity"]:
                                raise ValueError(f"课程 {course.code} 的名额不能低于当前已选教师人数")
                            for field, value in row.items():
                                setattr(course, field, value)
                            course.is_active = True
                            course.save()
                            updated += 1
                        else:
                            Course.objects.create(**row, is_active=True, created_by=request.user)
                            created += 1
            except ValueError as exc:
                messages.error(request, f"导入未执行：{exc}")
                return redirect("import_courses")
            request.session.pop(_pending_import_key("courses"), None)
            AuditLog.objects.create(user=request.user, action="导入课程", detail=f"新增 {created}，更新 {updated}")
            messages.success(request, f"导入完成：新增 {created} 门，更新 {updated} 门。")
            return redirect("course_list")
    form, preview = _preview_import(request, "courses", parse_course_import)
    return render(request, "selection/import_form.html", {"form": form, "kind": "课程", "preview": preview})


@portal_admin_required
def import_teachers(request):
    if request.method == "POST" and request.POST.get("action") == "confirm":
        pending = request.session.get(_pending_import_key("teachers"))
        if not pending:
            messages.error(request, "导入预览已失效，请重新选择文件。")
        else:
            created = updated = 0
            try:
                with transaction.atomic():
                    for row in pending["rows"]:
                        password_hash = row["password_hash"]
                        values = {key: value for key, value in row.items() if key != "password_hash"}
                        obj, was_created = User.objects.get_or_create(
                            username=row["username"], defaults={**values, "role": User.Role.TEACHER}
                        )
                        if obj.is_superuser or obj.role == User.Role.ADMIN:
                            raise ValueError(f"账号与管理员账号冲突：{obj.username}")
                        obj.display_name, obj.department, obj.role, obj.is_active = row["display_name"], row["department"], User.Role.TEACHER, True
                        if password_hash:
                            obj.password, obj.must_change_password = password_hash, True
                        obj.save()
                        created += was_created
                        updated += not was_created
            except ValueError as exc:
                messages.error(request, f"导入未执行：{exc}")
                return redirect("import_teachers")
            request.session.pop(_pending_import_key("teachers"), None)
            AuditLog.objects.create(user=request.user, action="导入教师", detail=f"新增 {created}，更新 {updated}")
            messages.success(request, f"导入完成：新增 {created} 人，更新 {updated} 人。")
            return redirect("teacher_list")
    form, preview = _preview_import(request, "teachers", parse_teacher_import)
    return render(request, "selection/import_form.html", {"form": form, "kind": "教师", "preview": preview})


def _workbook_response(workbook, filename):
    stream = BytesIO()
    workbook.save(stream)
    response = HttpResponse(
        stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={filename}"
    return response


def _style_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D4ED8")


@portal_admin_required
def download_template(request, kind):
    wb, ws = Workbook(), None
    ws = wb.active
    if kind == "teachers":
        ws.title = "教师账号"
        ws.append(["登录账号", "姓名", "部门", "初始密码"])
        ws.append(["teacher01", "张老师", "招生科", "ChangeMe123"])
        filename = "teachers-template.xlsx"
    else:
        ws.title = "课程"
        ws.append(["课程编号", "课程名称", "课程类别", "名额", "课程说明"])
        ws.append(["KC001", "就业指导", "通识课程", 10, "课程说明示例"])
        filename = "courses-template.xlsx"
    _style_header(ws)
    return _workbook_response(wb, filename)


@portal_admin_required
def results(request):
    query = request.GET.get("q", "").strip()
    department = request.GET.get("department", "").strip()
    selected = Selection.objects.select_related("teacher").order_by("selected_at")
    if department:
        selected = selected.filter(teacher__department=department)
    count_filter = Q(selections__teacher__department=department) if department else Q()
    courses = Course.objects.annotate(selection_total=Count("selections", filter=count_filter))
    if query:
        courses = courses.filter(Q(code__icontains=query) | Q(name__icontains=query))
    courses = courses.prefetch_related(Prefetch("selections", queryset=selected, to_attr="filtered_selections"))
    departments = User.objects.filter(role=User.Role.TEACHER).exclude(department="").order_by("department").values_list("department", flat=True).distinct()
    return render(request, "selection/results.html", {
        "courses": courses, "query": query, "department": department, "departments": departments,
    })


@portal_admin_required
def export_results(request):
    wb, ws = Workbook(), None
    ws = wb.active
    ws.title = "选课结果"
    ws.append(["课程编号", "课程名称", "姓名", "登录账号", "部门", "选课时间"])
    for item in Selection.objects.select_related("course", "teacher").order_by("course__code", "selected_at"):
        ws.append([
            item.course.code, item.course.name, item.teacher.display_name, item.teacher.username,
            item.teacher.department, item.selected_at.astimezone().strftime("%Y-%m-%d %H:%M:%S"),
        ])
    _style_header(ws)
    AuditLog.objects.create(user=request.user, action="导出选课结果")
    return _workbook_response(wb, "course-selection-results.xlsx")


@portal_admin_required
def export_unselected_teachers(request):
    department = request.GET.get("department", "").strip()
    teachers = User.objects.filter(role=User.Role.TEACHER, is_active=True, selections__isnull=True)
    if department:
        teachers = teachers.filter(department=department)
    wb, ws = Workbook(), None
    ws = wb.active
    ws.title = "未选课教师"
    ws.append(["姓名", "登录账号", "部门", "账号状态"])
    for teacher in teachers.order_by("department", "display_name", "username"):
        ws.append([teacher.display_name, teacher.username, teacher.department, "启用"])
    _style_header(ws)
    AuditLog.objects.create(user=request.user, action="导出未选课教师", detail=f"部门：{department or '全部'}")
    return _workbook_response(wb, "unselected-teachers.xlsx")


@teacher_required
def teacher_courses(request):
    query = request.GET.get("q", "").strip()
    courses = Course.objects.filter(is_active=True).annotate(selection_total=Count("selections"))
    if query:
        courses = courses.filter(Q(code__icontains=query) | Q(name__icontains=query) | Q(category__icontains=query))
    selected_ids = set(request.user.selections.values_list("course_id", flat=True))
    return render(request, "selection/teacher_courses.html", {
        "courses": courses, "selected_ids": selected_ids, "query": query,
        "selection_total": len(selected_ids), "setting": SiteSetting.load(),
    })


@teacher_required
@require_POST
def choose_course(request, pk):
    try:
        select_course(teacher=request.user, course_id=pk)
        messages.success(request, "选课成功。")
    except (ValidationError, Course.DoesNotExist) as exc:
        message = exc.messages[0] if isinstance(exc, ValidationError) else "课程不存在。"
        messages.error(request, message)
    return redirect("teacher_courses")


@teacher_required
@require_POST
def cancel_course(request, pk):
    try:
        cancel_selection(teacher=request.user, selection_id=pk)
        messages.success(request, "已取消该课程。")
    except ValidationError as exc:
        messages.error(request, exc.messages[0])
    return redirect("my_selections")


@teacher_required
def my_selections(request):
    items = request.user.selections.select_related("course")
    return render(request, "selection/my_selections.html", {"items": items, "setting": SiteSetting.load()})


def health(request):
    return HttpResponse("ok", content_type="text/plain")
